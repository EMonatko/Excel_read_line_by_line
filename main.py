def create_template(path_string) :
    """This function creates an excel file which will summerize all the data from other excels
       The input of is path string and the output is the date in format YMDHMS
    """
    today = datetime.now()
    today = today.strftime('%y%y%m%d%H%M%S')
    # print(today)
    temp_path = os.path.join(path_string, today)
    # temp_path = today
    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(f'{temp_path}.xlsx')
    worksheet0 = workbook.add_worksheet('ATR')  # Defaults to Sheet1.
    worksheet1 = workbook.add_worksheet('ESS')  # Data.
    worksheet2 = workbook.add_worksheet('Statistics')  # Defaults to Sheet

    # Some data we want to write to the worksheet.
    Tests_List = ['Temp', 'SN', 'Output Power @ P1dBCP', 'Output Power Control Range/Resolution, FWD PWR Ind',
                  'Output IP3', 'LO Carrier Leakage', 'Sideband Suppression',
                  'Frequency Accuracy and Stability', 'A1 - Noise Figure vs. Gain', 'A1 - Gain variability',
                  'A1 - Image Suppression vs. Gain', 'Spurious',
                  'A2 - Noise Figure vs. Gain', 'A2 - Gain variability', 'A2 - Image Suppression vs. Gain',
                  'Average Power Consumption', 'Input Voltage', 'Digital Tests'
                  ]

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    # col = 0

    # Iterate over the data and write it out row by row.
    for index in range(3) :
        for i in range(len(Tests_List)) :
            worksheet0.write(row, i, Tests_List[i])
            worksheet1.write(row, i, Tests_List[i])
            worksheet2.write(row, i, Tests_List[i])
            # col += 1

    workbook.close()

    return today, temp_path


def excel_fun_read(file_name, template_name, template_location, counter) :
    """
    Function receive 3 string inputs and searching for the 'PASS\ FAIL' section in the excel file
    :param file_name:
    :param template_name:
    :param template_location:
    :return:
    """
    for list_number in range(1, 4) :
        inputWorkbook = xlrd.open_workbook(file_name)
        inputWorksheet = inputWorkbook.sheet_by_index(list_number)
        rows = inputWorksheet.nrows
        cols = inputWorksheet.ncols
        print(f'{rows} Rows in the file\t')  # <- get rows number starts from 0
        print(f'{cols} Cols in the file\n')  # <- get coloms number starts from 0
        dictionary = {1 : 'ATR', 2 : 'ESS Hot cycle 1', 3 : 'ESS Cold cycle 1', 4 : 'ESS Hot cycle 2',
                      5 : 'ESS Cold cycle 2'}
        if cols == 9 :
            print('next file')
        if cols == 12 or cols == 9 :
            cols = 8
            sub = 2
        else :
            cols = 12
            sub = 3
        for excel_row in range(1, sub) :
            sn = int(inputWorksheet.cell_value(0, cols))
            print(f'working on 000{sn}.xlsx')  # <- Indicates which file is open
            TestLocation_list = []  # <- Creation of list
            PassFail_col_list = []  # <- Creation of list
            for i in range(rows) :
                # Follow the H colom check if there is 'PASS'/'FAIL' or empty cell
                # If empty cell skip it until the end of the excel file
                if inputWorksheet.cell_value(i, cols - excel_row) == 'PASS' or inputWorksheet.cell_value(i,
                                                                                                         cols - excel_row) == 'FAIL' or inputWorksheet.cell_value(
                        i, cols - excel_row) == 'N/T' :
                    TestLocation_list.append(i)
                    PassFail_col_list.append(str(inputWorksheet.cell_value(i, cols - excel_row)))

            location_list, len_of_every_test_list = Create_2_lists_of_locations(TestLocation_list, list_number)
            pass_fail_list = sort_list_of_pass_and_fail(len_of_every_test_list, location_list, file_name, list_number,
                                                        cols, PassFail_col_list, excel_row)

            # print(f'''It's the end of {list_number} in file 000{sn} excel_row = {excel_row}''')
            print(f'''It's the end of {dictionary.pop(counter)} in file 000{sn}\n''')
            write_to_excel(sn, template_location, pass_fail_list, counter)
            counter += 1

    print('''it's the end of the loop''')


def Create_2_lists_of_locations(TestLocation_list, list_number) :
    counter = 1
    temp_location_list = []
    temp_len_of_every_test_list = []
    for i_sub in range(len(TestLocation_list)) :
        try :
            if list_number == 1 :
                if TestLocation_list[i_sub] == 6 or TestLocation_list[i_sub] == 110 or TestLocation_list[
                    i_sub] == 139 or TestLocation_list[i_sub] == 172 or TestLocation_list[i_sub] == 224 :
                    temp_location_list.append(TestLocation_list[i_sub])
            if TestLocation_list[i_sub] + 1 == TestLocation_list[i_sub + 1] :
                if counter == 1 :
                    temp_location_list.append(TestLocation_list[i_sub])
                counter = counter + 1

            else :
                temp_len_of_every_test_list.append(counter)
                counter = 1
        except :
            pass
    return temp_location_list, temp_len_of_every_test_list


def sort_list_of_pass_and_fail(len_of_test_list, location_of_every_test_list, file_name, list_number, cols,
                               PassFail_col_list, excel_row) :
    Results_col = ['PASS']
    inputWorkbook = xlrd.open_workbook(file_name)
    inputWorksheet = inputWorkbook.sheet_by_index(list_number)
    try :
        for index in range(len(location_of_every_test_list)) :
            for i in range(location_of_every_test_list[index],
                           location_of_every_test_list[index] + len_of_test_list[index]) :
                if str(inputWorksheet.cell_value(i, cols - excel_row)) == 'FAIL' or str(
                        inputWorksheet.cell_value(i, cols - excel_row)) == 'N/T' :
                    Results_col[index] = str(inputWorksheet.cell_value(i, cols - excel_row))

            if len(Results_col) < len(location_of_every_test_list) :
                Results_col.append('PASS')
    except :
        pass
    print(Results_col)
    inputWorkbook.release_resources()
    return Results_col


def write_to_excel(sn, path, results, counter):
    temp_path = f'{path}.xlsx'
    workbook = openpyxl.load_workbook(filename=temp_path)
    print(workbook.sheetnames)
    atr = workbook['ATR']
    ess = workbook['ESS']
    temp_dictionary = {1 : 'Room', 2 : 'Hot', 3 : 'Cold'}
    temp = temp_dictionary.pop(1)
    if counter == 2 or counter == 4 :
        temp = temp_dictionary.pop(2)
    if counter == 3 or counter == 5 :
        temp = temp_dictionary.pop(3)
    sheet = workbook.active
    dictionary = {1 : atr, 2 : ess}
    print(sheet.title)
    if counter >= 2 :
        counter = 2
    dictionary_value = dictionary.pop(counter)

    for rows in range(1, 100) :
        if dictionary_value.cell(row=rows, column=1).value == None :
            print(f'empty row index: {rows}')

            dictionary_value.cell(row=rows, column=1, value=temp)
            dictionary_value.cell(row=rows, column=2, value=f'000{sn}')
            break

    for columns in range(3, len(results) + 3) :
        dictionary_value.cell(row=rows, column=columns, value=results[columns - 4])

    workbook.save(filename=temp_path)


def main(name) :
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.
    path_string = '/home/pi/Desktop/Python/fwdreportsnb'
    template_path_string = '/home/pi/Desktop/Pycharm/Tempalte'
    template_path_string = 'D:\Rasberry Pie\Python\Excel\Result'
    # path = input('Input Path location: \n')
    path_string = 'D:\Rasberry Pie\Python\Excel\Excel'
    excel_files_list = sorted([f for f in os.listdir(path_string) if f.endswith('.xlsx')])
    # excel_files = sorted(excel_files)
    print(excel_files_list, '\n')

    time_date, Template_path = create_template(template_path_string)
    print(time_date)

    for i in range(len(excel_files_list)) :
        counter = 1
        full_path = os.path.join(path_string, excel_files_list[i])
        print('\n', excel_files_list[i])
        excel_fun_read(full_path, time_date, Template_path, counter)


# Press the green button in the gutter to run the script.
if __name__ == '__main__' :
    import os
    import xlrd
    import xlsxwriter
    import openpyxl
    from xlwt import Workbook
    from datetime import datetime

    main('Script')
