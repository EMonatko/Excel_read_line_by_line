from openpyxl import Workbook
import openpyxl
import os
date = 20201124181754
template_path = 'D:\Rasberry Pie\Python\Excel\Result'
date, template_path = create_template(template_path)

file = os.path(template_path, date + '.xlsx')
#file = "enter_path_to_file_here"
wb = openpyxl.load_workbook(file)
ws = wb.active

for row in ws.iter_rows('A'):
    for cell in row:
        if cell.value == 'None':
            print(ws.cell(row=cell.row, column=2).value) #change column number for any cell value you want